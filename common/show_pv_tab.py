import pandas as pd
import numpy as np
import math
from importlib import reload
import common.dict_operator
reload(common.dict_operator)
from common.dict_operator import DPop, DSub, DSumMo
from data_process.local_sql import read_sql
from data_process.corr_analyze import excel_ws, get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, DataBar
from openpyxl.styles import NamedStyle, Font, Border, Side
side_none = Side(border_style=None)
bd_none = Border(top=side_none, left=side_none, right=side_none, bottom=side_none)

def get_lag_df(df, i, lag=10):
    if i >= 0:
        tmp_df = df.iloc[max(0, i - lag):(i + 1)]
    if i < 0:
        tmp_df = df.iloc[(i - lag):(i + 1)]
    return tmp_df

def show_pv(tmp_df):
    return pd.DataFrame(tmp_df.apply(lambda x:DPop(DSub(x["D_ask"], x["D_bid"])), axis=1).to_dict()).sort_index(ascending=False).fillna("")

def show_vol(tmp_df):
    tmp_df["VWAP"] = (tmp_df["amt"] / tmp_df["vol"]).round(2) - tmp_df["D_bid"]. apply(lambda x:max(x.keys()))
    tmp_df["vol"] = tmp_df["vol"]. astype(int)
    return tmp_df[["vol", "VWAP"]]. T

def show_exch_tab(tmp_df):
    return pd.DataFrame(tmp_df.apply(lambda x:DPop(x["D_exch"]), axis=1).to_dict()).sort_index(ascending=False).fillna("")

def split_tab(tmp_df):
    return pd.DataFrame("*", columns=tmp_df.index, index=[""])

def show3(df, i, lag=10):
    tmp_df = get_lag_df(df, i, lag)
    return pd.concat([show_pv(tmp_df),
           split_tab(tmp_df), 
           show_vol(tmp_df),
           split_tab(tmp_df),           
           show_exch_tab(tmp_df)]
          )

def show2(df, i, lag=10):
    tmp_df = get_lag_df(df, i, lag)
    return pd.concat([show_pv(tmp_df),
           split_tab(tmp_df), 
           show_exch_tab(tmp_df)]
          )



def show_excel(df, i, lag, path, tab="tab", mode='a'):
    tmp_df = get_lag_df(df, i, lag)
    df1 = show_pv(tmp_df)
    df2 = show_exch_tab(tmp_df)
    max_v1 = pd.Series([i for i in df1.values.flatten() if isinstance(i, (float, int))]).abs().quantile(0.95)
    max_v2 = pd.Series([i for i in df2.values.flatten() if isinstance(i, (float, int))]).abs().quantile(0.95)
    rule1 = ColorScaleRule(
        start_type='num',
        start_value= -max_v1,
        start_color='00FF0000',
        mid_type='num',
        mid_value=0,
        mid_color='00FFFFFF',
        end_type='num',
        end_value= max_v1,
        end_color='0000FF00'
    )        
    rule2 = DataBarRule(start_type='num',
                        start_value=0,
                        end_type='num',
                        end_value=max_v2,
                        color="FF638EC6",
                        showValue="None",
                        minLength=None,
                        maxLength=None)

    idx_max = max(df1.index[0], df2.index[0])
    idx_min = min(df1.index[ - 1], df2.index[ - 1])
    df3_idx = list(range(idx_max, idx_min - 1, -1))
    df3_col = list()
    for j, i in enumerate(df1.columns):
        df3_col.append(i)
        df3_col.append("")

    df3 = pd.DataFrame(index=df3_idx, columns=df3_col)
    for i in range(df1.shape[1]):
        df3.iloc[:, 2 * i + 1] = df1.iloc[:, i]
        df3.iloc[:, 2 * i] = df2.iloc[:, i]
        
    df3.fillna("", inplace=True)
    
    with pd.ExcelWriter(path, engine='openpyxl', mode=mode) as writer:
        df3.to_excel(writer, tab)
        ws = writer.book.get_sheet_by_name(tab)
        ws.__class__ = excel_ws
        ws.load_df(df3)    
        ws.cond_format_cols_list(cols=range(3, 2 + df3.shape[1], 2),
                                 start_row=2, 
                                 cond_format= rule1)
        ws.cond_format_cols_list(cols=range(2, 1 + df3.shape[1], 2),
                                 start_row=2,
                                 cond_format= rule2)

        for i in range(1, df3.shape[1] + 2):
            ws.column_dimensions[get_column_letter(i)].width = 6

        header_format = {
            "font":Font(size=10, name = "Courier New"),
            "border": bd_none,
        }
        ws.format_cols(None, None, header_format)
        ws.freeze_panes = ws['B2']
        
if __name__ == "__test__":
    tab = exch_detail("rb")
    tab.get_columns()
    df = read_sql(f"""select {','.join(tab.col_type["name"].tolist())} from rb.tickdata""")
    show_excel(df, 1500, 1500, "test1.xlsx")
    
    df[["date"]]
    read_sql("""select min(date) from rb.tickdata""")
