import pandas as pd
import numpy as np
import math
from common.append_df import cc1, cc2
import sys
from bin_tools.bins import binning, bins_simple_mean
import re
import gc
from talib_info import func_info
from talib_idx import table_talib_normal, talib_period
from kline import period_map, period_map_total

from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, coordinate_to_tuple
from openpyxl.utils import get_column_letter, get_column_interval
import openpyxl

side_none = Side(border_style=None)
bd_none = Border(top=side_none, left=side_none, right=side_none, bottom=side_none)
bd_up = Border(top=Side(border_style="medium"))

def cross_corr(dfx, dfy):
    dfx_v = (~dfx.isnull()).astype(np.float64)
    dfy_v = (~dfy.isnull()).astype(np.float64)
    dfx = dfx.fillna(0)
    dfy = dfy.fillna(0)

    corr_crosssum = dfy.T @ dfx
    corr_vcnt = (dfy_v.T) @ (dfx_v)
    corr_xsum = (dfy_v.T) @ (dfx)
    corr_ysum = (dfy.T) @ (dfx_v)

    # if fix_ymean:
    #     corr_ysum[:] = 0

    corr_xmean = (corr_xsum / corr_vcnt)
    corr_ymean = (corr_ysum / corr_vcnt)

    corr_xsqr = dfy_v.T @ (dfx**2)
    corr_ysqr = (dfy.T**2) @ (dfx_v)

    corr_xsqr_N = corr_xsqr - (corr_xmean**2) * (corr_vcnt)
    corr_ysqr_N = corr_ysqr - (corr_ymean**2) * (corr_vcnt)
    corr_crosssum_N = corr_crosssum - (corr_xmean * corr_ymean * corr_vcnt)
    corr_xy = corr_crosssum_N / ((corr_ysqr_N * corr_xsqr_N)**(1 / 2))
    return corr_xy

def daywise_corr(data, d, xidx, yidx):
    ls_d = list()
    ls_data = list()
    for i, x in data.groupby(d):
        ls_d.append(i)
        ls_data.append(cross_corr(x[xidx], x[yidx]))

    daywise_corr_df = np.array(ls_data)
    daywise_mean = pd.DataFrame(np.nanmean(daywise_corr_df, axis=0), index=yidx, columns=xidx)
    daywise_std = pd.DataFrame(np.nanstd(daywise_corr_df, axis=0), index=yidx, columns=xidx)
    daywise_sharpe = daywise_mean / daywise_std
    return daywise_sharpe, daywise_mean, daywise_std, daywise_corr_df

class excel_ws(openpyxl.worksheet.worksheet.Worksheet):
    def load_df(self, df):
        self.nrows = df.shape[0] + 1
        self.ncols = df.shape[1] + 1
        return self

    def get_row(self, i):
        return self[i:i]

    def get_col(self, i):
        j = get_column_letter(i)
        return self[f"{j}:{j}"]
    
    def format_cols(self, start, end, format_dict):
        if start is None:
            start = 1
        if end is None:
            end = self.ncols
        for i in range(start, end + 1):
            col = self.get_col(i)
            for cell in col:
                for j1, j2 in format_dict.items():
                    setattr(cell, j1, j2)

    def format_row(self, start, end, format_dict):
        if start is None:
            start = 1
        if end is None:
            end = self.nrows
        for i in range(start, end + 1):
            row = self.get_row(i)            
            for cell in row:
                for j1, j2 in format_dict.items():
                    setattr(cell, j1, j2)
    @staticmethod
    def format_str(i, j):
        return get_column_letter(j) + str(i)

    def cond_format(self, p0, p1, cond_format):
        s0 = self.format_str(p0[0], p0[1])
        s1 = self.format_str(p1[0], p1[1])
        srange = s0 + ":" + s1
        print(srange)
        self.conditional_formatting.add(srange, cond_format)

    def cond_format_cols(self, start, end, cond_format):
        if start is None:
            start = 1
        if end is None:
            end = self.ncols
        self.cond_format((1, start), (self.nrows, end), cond_format)
    
def beautify_excel_new(tmp_df,
                       sheet_name, 
                       writer,
                       conditional_format=None,
                       text_format=None,
                       header_format=None,
                       need_upborder=False, 
                       ):
    tmp_df.to_excel(writer, sheet_name=sheet_name)
    ws = writer.book.get_sheet_by_name(sheet_name)
    ws.__class__ = excel_ws
    ws.load_df(tmp_df)

    if conditional_format is None:
        conditional_format = ColorScaleRule(
            start_type='num',
            start_value= -0.1,
            start_color='00FF0000',
            mid_type='num',
            mid_value=0,
            mid_color='00FFFFFF',
            end_type='num',
            end_value= 0.1,
            end_color='0000FF00'
        )        
    if text_format is None:
        text_format = {
            "font":Font(size=10, name = "Courier New"),
            "number_format":"0.00",
        }        
    if header_format is None:
        header_format = {
            "font":Font(size=10, name = "Courier New"),
            "border": bd_none,
        }
    ws.format_cols(None, None, header_format)
    
    if isinstance(text_format, list):
        for _start, _end, _format in text_format:
            ws.format_cols(_start, _end, _format)
    else:
        ws.format_cols(1, ws.ncols, text_format)

    if isinstance(conditional_format, list):
        for _start, _end, _format in conditional_format:
            ws.cond_format_cols(_start, _end, _format)
    else:
        ws.cond_format_cols(1, ws.ncols, conditional_format)

    if need_upborder is True:
        first_rows = (pd.Series(tmp_df.index).drop_duplicates().index + 2).tolist()
        for i in first_rows:
            ws.format_row(i, i, {"border": bd_up})

    A_width = max(tmp_df.index.str.len())
    ws.column_dimensions["A"].width = A_width
    return ws

def beautify_excel(tmp_df,
                   sheet_name, 
                   writer, conditional_format=None,
                   text_format=None,
                   header_format=None,
                   ):
    tmp_df.to_excel(writer, sheet_name=sheet_name)
    if conditional_format is None:
        conditional_format = {
            'type': '3_color_scale',
            "min_type": "num",
            "max_type": "num",
            "mid_type": "num",                                          
            "min_value": "-0.1", 
            "mid_value": "0", 
            "max_value": "0.1", 
            "min_color": "red", 
            "mid_color": "white", 
            "max_color": "green", 
    }
    if text_format is None:
        text_format = {
            'num_format': '0.00%',
            "font": "Courier New",
            "font_size": 10,                         
        }
    if header_format is None:
        header_format = {
            "font": "Courier New",
            "font_size": 10, 
        }
    worksheet = writer.book.get_worksheet_by_name(sheet_name)

    header_format = writer.book.add_format(header_format)
    worksheet.set_column(0, tmp_df.shape[1], None, header_format)

    if isinstance(text_format, list):
        for _start, _end, _format in text_format:
            _format = writer.book.add_format(_format)
            worksheet.set_column(_start, _end, None, _format)
    else:
        _format = writer.book.add_format(text_format)        
        worksheet.set_column(0, tmp_df.shape[1], None, _format)

    if isinstance(conditional_format, list):
        for _start, _end, _format in conditional_format:
            worksheet.conditional_format(0, _start, tmp_df.shape[0], _end, _format)
    else:
        worksheet.conditional_format(0, 0, tmp_df.shape[0], tmp_df.shape[1], conditional_format)
        
    for i in range(tmp_df.shape[1]):
        worksheet.write(0, i + 1, str(tmp_df.columns[i]), header_format)

    for i in range(tmp_df.shape[0]):
        worksheet.write(i + 1, 0, str(tmp_df.index[i]), header_format)

    worksheet.write(0, 0, tmp_df.index.name, header_format)
    return worksheet

def get_idx_info(x:str):
    l = list()
    l1 = x.split("_")
    for j, i in enumerate(l1[:: -1]):
        try:
            l.append(float(i))
        except:
            break_point = j
            break
    tmp_str = "_". join(l1[:(len(l1) - len(l))])
    return tuple([tmp_str] + l[:: -1])

def get_idy_info(x:str):
    return period_map_total.get(x, 0)

def sort_index(tmp_df):
    tmp_df = tmp_df.loc[pd.Series({i:get_idx_info(i) for i in tmp_df.index}).sort_values().index]
    tmp_df = tmp_df[pd.Series({i:period_map_total.get(i, 0) for i in tmp_df.columns}).sort_values().index]
    return tmp_df
        
class xydata(pd.DataFrame):
    def __init__(self, data,
                 x_symbol="^TX", y_symbol="^ORM|^ORT", 
                 d = "date"
                 ):
        super().__init__(data)
        self.d = d
        self.idx = sorted(self.cc(x_symbol), key=get_idx_info)
        self.idy = sorted(self.cc(y_symbol), key=get_idy_info)
        self.dl = data[self.d]. drop_duplicates().sort_values()
    
    def erase_kline_tick_data(self, split_type=1):
        ### type 0 group by date
        ### type 1 group by (date (morning, afternoon,night))
        ### type 2 group by (date (day, night))
        erase_kline_tickdata(self, self.idx, self.idy, split_type=split_type)

    def cross_corr(self):
        self.cor = cross_corr(self[self.idx], self[self.idy])

    def daywise_corr(self):
        self.dcor_sharpe, self.dcor_mean, self.dcor_std, self.dcor_corr = \
            daywise_corr(self, self.d, self.idx, self.idy)

    @property
    def null_rate(self):
        return self.isnull().mean()[self.idx]

    def to_excel(self, path, append_info=None):
        with pd.ExcelWriter(path, engine='xlsxwriter') as writer:
            tmp_df = self.cor.T
            tmp_df = sort_index(tmp_df)
            beautify_excel(tmp_df, "corr", writer)
            for tmp_df, name in zip([self.dcor_sharpe.T,
                                     self.dcor_mean.T,
                                     self.dcor_std.T],
                                    ["dcorr_sharpee", "dcorr_mean", "dcorr_std"]):
                print(name)
                tmp_df = sort_index(tmp_df)
                v = tmp_df.melt()["value"]. abs()
                maximum = v[v < math.inf]. quantile(0.95)
                beautify_excel(tmp_df, name, writer,
                               conditional_format={
                                   'type': '3_color_scale',
                                   "min_type": "num",
                                   "max_type": "num",
                                   "mid_type": "num",                                          
                                   "min_value": f"{-maximum}", 
                                   "mid_value": "0", 
                                   "max_value": f"{maximum}", 
                                   "min_color": "red", 
                                   "mid_color": "white", 
                                   "max_color": "green", 
                               },
                               text_format = {
                                   'num_format': '0.00',
                                   "font": "Courier New",
                                   "font_size": 10,                         
                               })
                
            if append_info is not None:
                for i, j in append_info.items():
                    beautify_excel(j, i, writer,
                                   conditional_format={},
                                   text_format = {
                                       "font": "Courier New",
                                       "font_size": 10,                         
                                   })
                    
class talib_corr(table_talib_normal):
    def get_corr(self):
        self.data = xy_data(self. get_join_data())
        self.data.cross_corr()
        self.data.daywise_corr()

class corr_talib_basic(table_talib_normal):
    def __init__(self, code, period):
        super().__init__(code, f"kline_{period}", f"TXV_{period}")
        
    def get_corr(self, **params):
        self.data = xydata(self. get_join_data(), **params)
        self.data.cross_corr()
        self.data.daywise_corr()

class corr_talib_m(table_talib_normal):
    def __init__(self, code, period):
        super().__init__(code, f"kline_{period}", f"TXM_{period}")
        
    def get_corr(self, **params):
        self.data = xydata(self. get_join_data(), **params)
        self.data.cross_corr()
        self.data.daywise_corr()
        

def excel_3color(v=0.1):
    res = ColorScaleRule(
            start_type='num',
            start_value= v,
            start_color='00FF0000',
            mid_type='num',
            mid_value=0,
            mid_color='00FFFFFF',
            end_type='num',
            end_value= -v,
            end_color='0000FF00'
        )   
    return res

def excel_bar(color="#63C384"):
    res = DataBarRule(
        start_type='num',
        start_value=0,
        end_type='max',
        color="FF638EC6")    
    return res

class corr_pcluster(dict):
    #_period = talib_period[8:]
    _period = talib_period
    def __init__(self, code, cls, **params):
        super().__init__()
        for i in self._period:
            print(f"period {i}, fetch data ")
            self[i] = cls(code, i)
            print(f"period {i}, calc corr ")
            self[i]. get_corr(**params)
            # self[i]. data.cor["period"] = i
            # self[i]. data.dcor_sharpe["period"] = i
            # self[i]. data.dcor_mean["period"] = i
            # self[i]. data.dcor_std["period"] = i

    def null_rate(self):
        pass

    def lift_dim(self, i):
        return np.array([getattr(self[j1]. data, i) for j1 in sorted(self)])
            
    def lift_dict(self, i):
        obj = list(self.values())[0]. data
        idx = obj.idx
        idy = obj.idy      
        keys = sorted(self.keys())
        arr = self.lift_dim(i)                    
        if len(arr.shape) == 3:
            res = {idx[i]:arr[...,i] for i in range(arr.shape[ - 1])}            
            for i in res.keys():
                res[i] = pd.DataFrame(res[i])
                res[i].index = keys
                res[i].columns = idy
            return res

        if len(arr.shape) == 2:
            res = pd.DataFrame(arr)
            res.columns = idx
            res.index = keys
            return res

    def lift_concat(self, i):
        l = list()
        nr = self.lift_dict("null_rate")
        d = self.lift_dict(i)
        for _i in sorted(d.keys(), key=get_idx_info):
            _j = d[_i]
            _narate = nr[_i]
            _narate.name = "NArate"
            l.append(pd.concat([
                pd.Series(_i, name="name", index=_j.index),
                pd.Series(_j.index, name="Period", index=_j.index),
                _narate,
                _j[sorted(_j.columns, key=get_idy_info)]
            ], axis=1))
        return pd.concat(l).set_index("name")


    def to_excel(self, path, append_info=None):
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            for i in ["cor", "dcor_sharpe", "dcor_mean", "dcor_std"]:
            #for i in ["cor"]:                
                tmp_df = self.lift_concat(i)
                if i == "cor":
                    ws = beautify_excel_new(
                        tmp_df, i, writer,
                        conditional_format=[
                            [4, None, excel_3color()],
                            [3, 3, excel_bar()],             
                        ],
                        text_format=[
                            [3, None, {"number_format":"0.00%",}],
                        ],
                        need_upborder=True                        
                    )
                    ws.freeze_panes = ws['D2']
                else:
                    v = tmp_df.iloc[:, 2:].melt()["value"]. abs()
                    maximum = v[v < math.inf]. quantile(0.95)
                    ws = beautify_excel_new(
                        tmp_df, i, writer,
                        conditional_format=[
                            [4, None, excel_3color(maximum)],
                            [3, 3, excel_bar()],             
                        ],
                        text_format=[
                            [3, 3, {"number_format":"0.00%",}], 
                            [4, None, {"number_format":"0.00",}],
                        ],
                        need_upborder=True

                    )
                    ws.freeze_panes = ws['D2']
                    
            if append_info is not None:
                for i, j in append_info.items():
                    beautify_excel_new(j, i, writer,
                                       conditional_format=[])
                    
            
if __name__ == "__test__":
    # TXB_cluster[300]. data.idy.shape
    # self = TXB_cluster
    TXB_cluster = corr_pcluster("rb", corr_talib_basic, y_symbol="^ORM|^ORT|^OMRM|^OMRT")
    TXB_cluster.to_excel("./output/CORR_TXB.xlsx", append_info={"func_info": func_info})

    TXM_cluster = corr_pcluster("rb", corr_talib_m, y_symbol="^ORM|^ORT|^OMRM|^OMRT")
    TXM_cluster.to_excel("./output/CORR_TXM.xlsx", append_info={"func_info": func_info})
    
    
    TXB_cluster.to_excel = MethodType(to_excel, TXB_cluster)
    from types import MethodType
    from talib.abstract import * 
    STOCHF(TXB_cluster[1]. data)["fastd"]. isnull().mean()
    (STOCHF(TXB_cluster[1]. data)["fastk"]. isnull()).mean()
    STOCHF(TXB_cluster[1]. data).iloc[1105:1205]
    STOCHF(TXB_cluster[1]. data)[STOCHF(TXB_cluster[1]. data)4["fastk"]. isnull()]. head(200)
    TXB_cluster[1]. data[~STOCHF(TXB_cluster[1]. data)["fastd"]. isnull()]. shape    
    





    
